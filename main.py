import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox
from PyQt5 import uic, QtWidgets
from datetime import datetime
from PyQt5.QtGui import QFont as QF
import pandas as pd
from openpyxl import load_workbook


def informe_sensibilizacion(namefile, nuevo_nombre):
    file_name = namefile
    df = pd.read_excel(file_name)  # , sheet_name= my_sheet)
    valor = df.drop("Nombre de usuario", axis=1)
    valor = valor.drop("Institución", axis=1)
    valor = valor.drop("Departamento", axis=1)
    valor = valor.drop("Dirección de correo", axis=1)
    valor = valor.drop("Última descarga de este curso", axis=1)
    valor = valor.replace({"-": 0})

    Actividad1 = valor['Tarea:Actividad 1. Revisar y ajustar el perfil (Real)']
    Actividad1[Actividad1 > 1] = 1
    Actividad1 = Actividad1.value_counts()

    Actividad2 = valor[
        'Foro:Rating grade for Actividad 2. Foro de presentación (Real)']  # Foro:Actividad 2. Foro de presentación (Real)
    Actividad2[Actividad2 > 1] = 1
    Actividad2 = Actividad2.value_counts()

    Actividad3 = valor['Cuestionario:Actividad 3. Tutoría virtual con Ude@ (Real)']
    Actividad3[Actividad3 > 1] = 1
    Actividad3 = Actividad3.value_counts()

    Actividad4 = valor['Tarea:Actividad 4. Planeación del espacio y tiempo en la virtualidad (Real)']
    Actividad4[Actividad4 > 1] = 1
    Actividad4 = Actividad4.value_counts()

    Actividad5 = valor['Tarea:Reporte de calificación actividad 5. Utilizando la caja de herramientas (Real)']
    Actividad5[Actividad5 > 1] = 1
    Actividad5 = Actividad5.value_counts()

    total = valor.shape[0]
    Actividad6 = valor['Cuestionario:Actividad 6. Evaluación final (Real)']
    Actividad6[Actividad6 > 1] = 1
    Actividad6 = Actividad6.value_counts()

    Reporte = valor['Total del curso (Real)']
    score = Reporte.value_counts()
    participacion = str(total - score[0])
    Reporte[Reporte < 70] = 0
    Reporte[Reporte >= 70] = 1

    Reporte = Reporte.value_counts()

    valor = valor.rename(columns={'Tarea:Actividad 1. Revisar y ajustar el perfil (Real)': 'Actividad 1',
                                  'Foro:Rating grade for Actividad 2. Foro de presentación (Real)': 'Actividad2',
                                  'Cuestionario:Actividad 3. Tutoría virtual con Ude@ (Real)': 'Actividad3',
                                  'Tarea:Actividad 4. Planeación del espacio y tiempo en la virtualidad (Real)': 'Actividad4',
                                  'Tarea:Reporte de calificación actividad 5. Utilizando la caja de herramientas (Real)': 'Actividad5',
                                  'Cuestionario:Actividad 6. Evaluación final (Real)': 'Actividad6',
                                  'Total del curso (Real)': 'Estado'})

    file = nuevo_nombre
    valor.to_excel(file, sheet_name='Informe')

    book = load_workbook(file)
    writer = pd.ExcelWriter(file, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    total = valor.shape[0]
    actividad_data_frame = {
        'Actividad': ['Actividad1', 'Actividad2', ' Actividad3', 'Actividad4', 'Actividad5', 'Actividad6'],
        'Participación': [total - Actividad1[0], total - Actividad2[0], total - Actividad3[0], total - Actividad4[0],
                          total - Actividad5[0], total - Actividad6[0]]}

    df1 = pd.DataFrame(actividad_data_frame)
    df1.to_excel(writer, "Hoja 2")

    aprobado_data = {'Aprobado': ['No aprobado', 'Participación'], str(total - Reporte[0]): [Reporte[0], participacion]}

    df2 = pd.DataFrame(aprobado_data)
    df2.to_excel(writer, "hoja 3")
    writer.save()


def informe_fpd(namefile, new_name):
    file_name = namefile
    df = pd.read_excel(file_name)  # , sheet_name= my_sheet)
    valor = df.drop("Nombre de usuario", axis=1)
    valor = valor.drop("Institución", axis=1)
    valor = valor.drop("Departamento", axis=1)
    valor = valor.drop("Dirección de correo", axis=1)
    valor = valor.drop("Última descarga de este curso", axis=1)
    valor = valor.replace({"-": 0})

    Actividad1 = valor['Tarea:Actividad 1. Revisar y ajustar el perfil (Real)']
    Actividad1[Actividad1 == "Aprobado"] = 1
    Actividad1[Actividad1 == "No aprobado"] = 0
    Actividad1 = Actividad1.value_counts()

    Actividad2 = valor['Foro:Rating grade for Actividad 2. Foro de presentación (Real)']
    Actividad2[Actividad2 == "Aprobado"] = 1
    Actividad2[Actividad2 == "No aprobado"] = 0
    Actividad2 = Actividad2.value_counts()

    Actividad3 = valor['Cuestionario:Actividad 3. Tutoría virtual con Ude@ (Real)']
    Actividad3[Actividad3 > 0] = 1
    Actividad3 = Actividad3.value_counts()

    Actividad4 = valor['Tarea:Actividad 4. Planeación del espacio y tiempo en la virtualidad (Real)']
    Actividad4[Actividad4 == "Aprobado"] = 1
    Actividad4[Actividad4 == "No aprobado"] = 0
    Actividad4 = Actividad4.value_counts()

    Actividad5 = valor['Cuestionario:Actividad 5. Evaluación final (Real)']
    Actividad5[Actividad5 > 0] = 1
    Actividad5 = Actividad5.value_counts()

    Total_curso = valor['Total del curso (Real)']
    Total_curso = Total_curso.value_counts()

    valor = valor.rename(columns={'Tarea:Actividad 1. Revisar y ajustar el perfil (Real)': 'Actividad 1',
                                  'Foro:Rating grade for Actividad 2. Foro de presentación (Real)': 'Actividad 2',
                                  'Cuestionario:Actividad 3. Tutoría virtual con Ude@ (Real)': 'Actividad 3',
                                  'Tarea:Actividad 4. Planeación del espacio y tiempo en la virtualidad (Real)': 'Actividad 4',
                                  'Cuestionario:Actividad 5. Evaluación final (Real)': 'Actividad 5',
                                  'Total del curso (Real)': 'Total del curso'

                                  })

    file = new_name
    valor.to_excel(file, sheet_name='Calificaciones')

    book = load_workbook(file)
    writer = pd.ExcelWriter(file, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    total = valor.shape[0]
    actividad_data_frame = {'Actividad': ['Actividad1', 'Actividad2', ' Actividad3', 'Actividad4', 'Actividad5'],
                            'Participación': [total - Actividad1[0], total - Actividad2[0], total - Actividad3[0],
                                              total - Actividad4[0], total - Actividad5[0]]}

    df1 = pd.DataFrame(actividad_data_frame)
    df1.to_excel(writer, "Hoja 2")

    aprobado_data = {'Participaron': ['Matriculados'], str(total - Total_curso[0]): total}

    df2 = pd.DataFrame(aprobado_data)
    df2.to_excel(writer, "hoja 3")
    writer.save()


class Ui_MainWindow(QMainWindow):
    ruta = ""

    def __init__(self):
        QMainWindow.__init__(self)
        uic.loadUi("Ventanas/mainwindow.ui", self)
        self.setWindowTitle("Informes de Cursos de Con Tic Aprendo")
        self.bt_seleccionar.clicked.connect(self.seleccionar)
        self.bt_sensibilizacion.clicked.connect(self.sensibilizacion)
        self.bt_fpd.clicked.connect(self.fpd)

    def showEvent(self, event):
        fecha = datetime.now().strftime("%d-%m-%Y")
        self.lb_fecha.setText(fecha)
        self.lb_fecha.setFont(QF("Sanserif ", 16))
        self.lb_fecha.setStyleSheet('color:green')

    def seleccionar(self):
        file = QtWidgets.QFileDialog.getOpenFileName(self, 'C:/Users/jdavi/Downloads/', self.ruta)
        if file:
            self.ruta = file[0]
            self.lb_ruta.setText(str(self.ruta))

    def sensibilizacion(self):
        new_name = self.le_name.text()
        if new_name and self.ruta:
            new_name = new_name + ".xlsx"
            informe_sensibilizacion(self.ruta, new_name)

        else:
            QMessageBox.warning(None, "Nombre", "Faltan datos", QMessageBox.Ok)

    def fpd(self):
        new_name = self.le_name.text()
        if new_name and self.ruta:
            new_name = new_name + ".xlsx"
            informe_fpd(self.ruta, new_name)
        else:
            QMessageBox.warning(None, "Datos", "Faltan datos", QMessageBox.Ok)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    principal = Ui_MainWindow()
    principal.show()
    app.exec_()
